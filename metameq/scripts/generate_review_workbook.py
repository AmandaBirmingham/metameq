"""Generate an Excel workbook for human review of golden type-pair outputs.

Reads the golden CSV files from expected_type_pair_outputs/ and produces a
three-tab Excel workbook that highlights meaningful values, groups by host
type, and makes it easy for a domain expert to verify correctness.

Usage:
    python -m metameq.scripts.generate_review_workbook [--golden-dir DIR] [--output PATH]
"""

import argparse
import glob
import os
from collections import defaultdict

import pandas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_GOLDEN_DIR = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "tests", "data", "expected_type_pair_outputs")
DEFAULT_OUTPUT = "type_pair_review.xlsx"

# Values treated as "needs user input" rather than meaningful pre-set data
DEFAULT_VALUES = {"not provided", "not applicable"}

# Column excluded from review (it's just the test identifier)
SKIP_COLUMNS = {"sample_name"}

# Fields that just echo the sample type name — only show if value differs
SAMPLE_TYPE_ECHO_FIELDS = {"sample_type", "qiita_sample_type"}

# Field used for grouping sample types within a host on Tab 2
GROUPING_FIELD = "env_package"

# Styles — headers are italic+bold to distinguish from normal field names
HOST_HEADER_FONT = Font(bold=True, size=14, color="FFFFFF")
HOST_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4",
                               fill_type="solid")
SECTION_HEADER_FONT = Font(bold=True, italic=True, size=11)
GROUP_HEADER_FONT = Font(bold=True, italic=True, size=11, color="2E75B6")
SAMPLE_HEADER_FONT = Font(bold=True, size=12)
FIELD_FONT = Font()
FIELD_FONT_BOLD = Font(bold=True)
NORMAL_FONT = Font()
DEFAULT_HIGHLIGHT_FILL = PatternFill(
    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def load_golden_files(golden_dir):
    """Load all golden CSVs and return a dict keyed by (host_type, sample_type)."""
    files = sorted(glob.glob(os.path.join(golden_dir, "*.csv")))
    data = {}
    for fp in files:
        name = os.path.splitext(os.path.basename(fp))[0]
        parts = name.split("__", 1)
        if len(parts) != 2:
            continue
        host_type, sample_type = parts
        df = pandas.read_csv(fp, dtype=str, keep_default_na=False)
        row = {}
        for col in df.columns:
            if col not in SKIP_COLUMNS:
                row[col] = df[col].iloc[0]
        data[(host_type, sample_type)] = row
    return data


def group_by_host(data):
    """Group (host_type, sample_type) -> row_dict into host_type -> {sample_type: row_dict}."""
    groups = defaultdict(dict)
    for (host_type, sample_type), row in data.items():
        groups[host_type][sample_type] = row
    return dict(sorted(groups.items()))


def compute_host_base(sample_rows):
    """Compute the base (shared) fields for a host type.

    Returns:
        base_meaningful: dict of {field: value} for shared non-default values
        base_default: dict of {field: default_value} for shared default values
    """
    sample_types = sorted(sample_rows.keys())
    if not sample_types:
        return {}, {}

    all_col_sets = [set(sample_rows[st].keys()) for st in sample_types]
    shared_cols = sorted(set.intersection(*all_col_sets))

    base_meaningful = {}
    base_default = {}

    for col in shared_cols:
        values = {sample_rows[st][col] for st in sample_types}
        if len(values) == 1:
            val = values.pop()
            if val in DEFAULT_VALUES:
                base_default[col] = val
            else:
                base_meaningful[col] = val

    return base_meaningful, base_default


def compute_overrides(sample_rows, base_meaningful, base_default):
    """Compute per-sample-type overrides relative to the host base.

    Filters out sample_type/qiita_sample_type when their value just echoes
    the sample type name (redundant info).

    Returns:
        dict of {sample_type: {
            "preset": {field: value},
            "user_input": {field: default},
        }}
    """
    base_all = {}
    base_all.update(base_meaningful)
    base_all.update(base_default)

    overrides = {}
    for sample_type in sorted(sample_rows.keys()):
        row = sample_rows[sample_type]
        preset = {}
        user_input = {}
        for col, val in sorted(row.items()):
            if col in base_all and base_all[col] == val:
                continue  # same as base, skip
            # Skip sample_type echo fields unless value differs from the sample type name
            if col in SAMPLE_TYPE_ECHO_FIELDS and val == sample_type:
                continue
            if val in DEFAULT_VALUES:
                user_input[col] = val
            else:
                preset[col] = val
        overrides[sample_type] = {"preset": preset, "user_input": user_input}
    return overrides


def compute_varying_columns(sample_rows, base_meaningful, base_default):
    """Find columns that vary across sample types (for the comparison matrix)."""
    base_all = set(base_meaningful.keys()) | set(base_default.keys())
    all_cols = set()
    for row in sample_rows.values():
        all_cols.update(row.keys())
    varying = sorted(all_cols - base_all)
    return varying


def group_sample_types_by_field(sample_rows, overrides, field_name):
    """Group sample types by the value of a specific field.

    Returns an ordered dict of {group_label: [sample_type, ...]} sorted by group label.
    """
    groups = defaultdict(list)
    for sample_type in sorted(sample_rows.keys()):
        row = sample_rows[sample_type]
        group_val = row.get(field_name, "other")
        groups[group_val].append(sample_type)
    return dict(sorted(groups.items()))


def auto_width(ws, min_width=10, max_width=50):
    """Auto-size column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[col_letter].width = width


def _write_host_header(ws, row_num, host_type, sample_count, num_cols=2):
    """Write a blue host type header row."""
    ws.cell(row=row_num, column=1, value=f"Host type: {host_type}")
    ws.cell(row=row_num, column=1).font = HOST_HEADER_FONT
    for c in range(1, num_cols + 1):
        ws.cell(row=row_num, column=c).fill = HOST_HEADER_FILL
    if sample_count is not None:
        ws.cell(row=row_num, column=2,
                value=f"({sample_count} sample type{'s' if sample_count != 1 else ''})")
        ws.cell(row=row_num, column=2).font = HOST_HEADER_FONT


def write_host_bases_tab(wb, host_groups):
    """Write Tab 1: Host Type Bases."""
    ws = wb.active
    ws.title = "Host Type Bases"
    row_num = 1

    for host_type, sample_rows in host_groups.items():
        base_meaningful, base_default = compute_host_base(sample_rows)

        _write_host_header(ws, row_num, host_type, len(sample_rows))
        row_num += 1

        # Sub-section: Pre-set values
        ws.cell(row=row_num, column=1, value="Pre-set values")
        ws.cell(row=row_num, column=1).font = SECTION_HEADER_FONT
        row_num += 1

        if base_meaningful:
            for field in sorted(base_meaningful.keys()):
                ws.cell(row=row_num, column=1, value=field)
                ws.cell(row=row_num, column=2, value=base_meaningful[field])
                row_num += 1
        else:
            ws.cell(row=row_num, column=1, value="(none)")
            row_num += 1

        row_num += 1  # blank row

        # Sub-section: Fields requiring user input, grouped by default value
        defaults_by_value = defaultdict(list)
        for field, val in base_default.items():
            defaults_by_value[val].append(field)

        for default_val in sorted(defaults_by_value.keys()):
            fields = sorted(defaults_by_value[default_val])
            ws.cell(row=row_num, column=1,
                    value=f"Fields requiring user input (default: '{default_val}')")
            ws.cell(row=row_num, column=1).font = SECTION_HEADER_FONT
            row_num += 1
            for field in fields:
                ws.cell(row=row_num, column=1, value=field)
                row_num += 1

        row_num += 2  # extra blank rows between host types

    auto_width(ws)


def _write_override_block(ws, row_num, col_offset, sample_type, override):
    """Write a single sample type's override block at the given column offset.

    Returns the number of rows written.
    """
    start_row = row_num
    preset = override["preset"]
    user_input = override["user_input"]

    # Sample type header
    ws.cell(row=row_num, column=col_offset, value=sample_type)
    ws.cell(row=row_num, column=col_offset).font = SAMPLE_HEADER_FONT
    row_num += 1

    if preset:
        ws.cell(row=row_num, column=col_offset, value="Pre-set values")
        ws.cell(row=row_num, column=col_offset).font = SECTION_HEADER_FONT
        row_num += 1
        for field in sorted(preset.keys()):
            ws.cell(row=row_num, column=col_offset, value=field)
            ws.cell(row=row_num, column=col_offset + 1, value=preset[field])
            row_num += 1

    if user_input:
        by_default = defaultdict(list)
        for field, val in user_input.items():
            by_default[val].append(field)
        for default_val in sorted(by_default.keys()):
            fields = sorted(by_default[default_val])
            ws.cell(row=row_num, column=col_offset,
                    value=f"Additional user input (default: '{default_val}')")
            ws.cell(row=row_num, column=col_offset).font = SECTION_HEADER_FONT
            row_num += 1
            for field in fields:
                ws.cell(row=row_num, column=col_offset, value=field)
                row_num += 1

    if not preset and not user_input:
        ws.cell(row=row_num, column=col_offset,
                value="(identical to host base)")
        ws.cell(row=row_num, column=col_offset).font = Font(italic=True)
        row_num += 1

    return row_num - start_row


def write_per_pair_overrides_tab(wb, host_groups):
    """Write Tab 2: Per-Pair Overrides, grouped by env_package with side-by-side layout."""
    ws = wb.create_sheet("Per-Pair Overrides")
    row_num = 1

    for host_type, sample_rows in host_groups.items():
        base_meaningful, base_default = compute_host_base(sample_rows)
        overrides = compute_overrides(sample_rows, base_meaningful, base_default)

        # Check if this host type has any overrides at all
        has_any_overrides = any(
            override["preset"] or override["user_input"]
            for override in overrides.values())
        if not has_any_overrides:
            continue  # skip host types where all pairs are identical to base

        # Group sample types by env_package
        env_groups = group_sample_types_by_field(
            sample_rows, overrides, GROUPING_FIELD)

        # Count how many groups have overrides to determine header width
        groups_with_overrides = [
            (label, sts) for label, sts in env_groups.items()
            if any(overrides[st]["preset"] or overrides[st]["user_input"]
                   for st in sts)]
        # Each group uses 3 columns (field, value, gap); last group needs only 2
        header_cols = max(2, len(groups_with_overrides) * 3 - 1)

        # Host type header spanning all column groups
        _write_host_header(ws, row_num, host_type, len(sample_rows),
                           num_cols=header_cols)
        row_num += 1

        # Lay out each env_package group as side-by-side column pairs
        col_offset = 1
        group_start_row = row_num

        for group_label, group_sample_types in groups_with_overrides:
            # Group header
            current_row = group_start_row
            ws.cell(row=current_row, column=col_offset, value=group_label)
            ws.cell(row=current_row, column=col_offset).font = GROUP_HEADER_FONT
            current_row += 1

            for sample_type in group_sample_types:
                override = overrides[sample_type]
                if not override["preset"] and not override["user_input"]:
                    continue  # skip sample types with no overrides

                rows_written = _write_override_block(
                    ws, current_row, col_offset, sample_type, override)
                current_row += rows_written + 1  # +1 for blank row between samples

            # Track the max row used across all groups
            row_num = max(row_num, current_row)

            # Move to next column pair (with a blank column gap)
            col_offset += 3

        row_num += 2  # extra blank rows between host types

    auto_width(ws)


def write_comparison_matrix_tab(wb, host_groups):
    """Write Tab 3: Comparison Matrix."""
    ws = wb.create_sheet("Comparison Matrix")
    row_num = 1

    for host_type, sample_rows in host_groups.items():
        base_meaningful, base_default = compute_host_base(sample_rows)
        varying_cols = compute_varying_columns(
            sample_rows, base_meaningful, base_default)

        if not varying_cols:
            ws.cell(row=row_num, column=1, value=f"Host type: {host_type}")
            ws.cell(row=row_num, column=1).font = HOST_HEADER_FONT
            ws.cell(row=row_num, column=1).fill = HOST_HEADER_FILL
            row_num += 1
            ws.cell(row=row_num, column=1,
                    value="(no varying columns across sample types)")
            row_num += 2
            continue

        _write_host_header(ws, row_num, host_type, len(sample_rows),
                           num_cols=len(varying_cols) + 1)
        row_num += 1

        # Column headers
        ws.cell(row=row_num, column=1, value="sample_type")
        ws.cell(row=row_num, column=1).font = FIELD_FONT_BOLD
        for col_idx, col_name in enumerate(varying_cols, start=2):
            ws.cell(row=row_num, column=col_idx, value=col_name)
            ws.cell(row=row_num, column=col_idx).font = FIELD_FONT_BOLD
        row_num += 1

        # Data rows
        for sample_type in sorted(sample_rows.keys()):
            row = sample_rows[sample_type]
            ws.cell(row=row_num, column=1, value=sample_type)
            ws.cell(row=row_num, column=1).font = FIELD_FONT_BOLD
            for col_idx, col_name in enumerate(varying_cols, start=2):
                val = row.get(col_name, "")
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                if val in DEFAULT_VALUES:
                    cell.fill = DEFAULT_HIGHLIGHT_FILL
            row_num += 1

        row_num += 2

    auto_width(ws)


def generate_workbook(golden_dir, output_path):
    """Generate the review workbook from golden files."""
    data = load_golden_files(golden_dir)
    host_groups = group_by_host(data)

    print(f"Loaded {len(data)} golden files across {len(host_groups)} host types")

    wb = Workbook()
    write_host_bases_tab(wb, host_groups)
    write_per_pair_overrides_tab(wb, host_groups)
    write_comparison_matrix_tab(wb, host_groups)
    wb.save(output_path)

    print(f"Wrote review workbook to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Excel review workbook from golden type-pair outputs.")
    parser.add_argument(
        "--golden-dir", default=DEFAULT_GOLDEN_DIR,
        help="Directory containing golden CSV files "
             f"(default: {DEFAULT_GOLDEN_DIR})")
    parser.add_argument(
        "--output", default=DEFAULT_OUTPUT,
        help=f"Output Excel file path (default: {DEFAULT_OUTPUT})")
    args = parser.parse_args()

    generate_workbook(args.golden_dir, args.output)


if __name__ == "__main__":
    main()
